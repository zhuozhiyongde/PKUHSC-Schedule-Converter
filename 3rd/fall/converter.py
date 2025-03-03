# -*- encoding:utf-8 -*-
# @Time		:	2022/08/30 02:01:39
# @File		:	converter.py
# @Author	:	Arthals
# @Contact	:	zhuozhiyongde@126.com
# @Software	:	Visual Studio Code

from openpyxl import Workbook, load_workbook
import os
import re

# wb = Workbook()
# ws = wb.create_sheet('mysheet', 0)
# wb.save('test.xlsx')
# wb.close()

wb = load_workbook("data.xlsx")
ws = wb["sheet1"]

maxRow = ws.max_row
maxCol = ws.max_column


def extractInterger(strin):
    return int(re.findall(r"\d+", strin)[0])


def extractWeek(strin):
    strin = re.sub(r"[周()]", "", strin)
    weeks = re.sub(r",", r"、", strin)
    return weeks


def extractDay(strin):
    dayDic = {
        "星期一": 1,
        "星期二": 2,
        "星期三": 3,
        "星期四": 4,
        "星期五": 5,
        "星期六": 6,
        "星期日": 7,
    }
    return dayDic[strin]


courseList = []
for row in range(2, maxRow + 1):
    courseName = ws.cell(row=row, column=2).value
    courseStart = extractInterger(ws.cell(row=row, column=8).value)
    courseEnd = extractInterger(ws.cell(row=row, column=9).value)
    courseWeek = extractWeek(ws.cell(row=row, column=6).value)
    courseDay = extractDay(ws.cell(row=row, column=7).value)
    courseLocation = ws.cell(row=row, column=11).value
    courseTeacher = ws.cell(row=row, column=10).value
    courseList.append(
        [
            courseName,
            courseDay,
            courseStart,
            courseEnd,
            courseTeacher,
            courseLocation,
            courseWeek,
        ]
    )

# print(courseList)

wb.close()

output = open("mySchedule.csv", "w+")
output.write("课程名称,星期,开始节数,结束节数,老师,地点,周数\n")
for course in courseList:
    for info in range(len(course)):
        course[info] = f'"{course[info]}"'
    print(course)
    output.write(",".join("%s" % id for id in (course)) + "\n")
output.close()
