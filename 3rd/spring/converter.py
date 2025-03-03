#!/usr/bin/env python
# -*- encoding: utf-8 -*-
# @Author  :   Arthals
# @File    :   converter.py
# @Time    :   2024/02/20 09:19:31
# @Contact :   zhuozhiyongde@126.com
# @Software:   Visual Studio Code

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
import os


def week_to_number(week_str):
    # print(week_str)
    if not re.match(r"周", week_str):
        week_str = "周" + week_str
    week_dict = {
        "周一": 1,
        "周二": 2,
        "周三": 3,
        "周四": 4,
        "周五": 5,
        "周六": 6,
        "周日": 7,
    }
    return week_dict.get(week_str, "输入错误")


def convert_theroy(sheet, output_file):
    def calculate_week(date_str):
        # 假设是闰年2024年
        year_str = "2024"

        # 将年份信息加入到日期字符串中
        start_date_str = year_str + ".2.19"
        current_date_str = year_str + "." + date_str

        # 使用包含年份的格式进行解析
        start_date = datetime.strptime(start_date_str, "%Y.%m.%d")
        current_date = datetime.strptime(current_date_str, "%Y.%m.%d")

        # 计算当前日期与第一周周一的差
        delta = current_date - start_date

        # 计算周数，由于delta.days是从0开始计算的，所以要加1
        # 然后除以7得到周数，由于是从周一开始算，所以需要加1
        week_number = delta.days // 7 + 1

        return week_number

    # 读取 data.xlsx 文件，加载其中的 “体检理论” 工作表，不含标题行
    df = pd.read_excel("data.xlsx", sheet_name=sheet, header=None)
    print("正在处理", sheet, "工作表", "...")
    # 打印行数，列数
    # print(df.shape)

    # 逐行检查，如果一行内只有一个单元格横跨所有列，那么这一行是一个标题行，打印
    for i in range(df.shape[0]):
        if i == 0:
            subject = (
                df.iloc[i]
                .values[0]
                .replace("2021级北医预防医学专业", "")
                .replace("教学课程表", "")
            )
            print(subject)
        if df.iloc[i].count() == 1:
            # print(df.iloc[i].values[0])
            continue
        else:
            # 打印非标题行的内容
            print(df.iloc[i].values)
            # 获取其中的内容为 星期，节次，一线教师，授课地点，日期 所在的列索引
            week_index = df.iloc[i].tolist().index("星期")
            section_index = df.iloc[i].tolist().index("节次")
            teacher_index = df.iloc[i].tolist().index("一线教师")
            place_index = df.iloc[i].tolist().index("授课地点")
            date_index = df.iloc[i].tolist().index("日期")
            # content_index 为 课程内容 或 见习内容 所在的列索引
            try:
                content_index = df.iloc[i].tolist().index("授课内容")
            except:
                content_index = df.iloc[i].tolist().index("见习内容")
            cur_index = i + 1
            break
    while True:
        is_exam = False
        # 检查是否又有标题行
        if df.iloc[cur_index].count() == 1:
            print(df.iloc[cur_index].values[0])
            break

        # 获取星期，节次，一线教师，授课地点，日期
        week = df.iloc[cur_index].tolist()[week_index]
        section = df.iloc[cur_index].tolist()[section_index]
        teacher = df.iloc[cur_index].tolist()[teacher_index]
        place = df.iloc[cur_index].tolist()[place_index].replace("\n", "")
        date = df.iloc[cur_index].tolist()[date_index]
        if "-" in section:
            start_section, end_section = section.split("-")
        else:
            start_section = end_section = section

        # 写入到 csv 文件
        # 课程名称,星期,开始节数,结束节数,老师,地点,周数
        # 首先计算周数
        # 已知日期格式为  3.4
        # 2.19 为第一周的周一
        # 计算周数
        week_number = calculate_week(date)
        # 写入到 csv 文件
        # 如果 content_index 内容含有 “考试” 字样：
        content = df.iloc[cur_index].tolist()[content_index]
        if "考试" in content:
            if section == "上午":
                start_section = "1"
                end_section = "4"
            else:
                start_section = "5"
                end_section = "8"
            is_exam = True
            teacher = "希波克拉底"

        with open(output_file, "a") as f:
            name = subject
            if is_exam:
                name = content
            # print([week])
            week = week_to_number(week)
            teacher = teacher.replace("\n", "")
            f.write(
                f"{name},{week},{start_section},{end_section},{teacher},{place},{week_number}\n"
            )
        cur_index += 1


def convert_internship(sheet, group, output_file):
    def calculate_time(date):
        date_str, section, week = date.split("/")
        # 定义第一周周一的日期
        year_str = "2024."
        start_date = datetime.strptime(year_str + "2.19", "%Y.%m.%d")
        # 将输入的日期字符串转换为日期格式
        current_date = datetime.strptime(year_str + date_str, "%Y.%m.%d")

        # 计算当前日期与第一周周一的差
        delta = current_date - start_date

        # 计算周数，由于delta.days是从0开始计算的，所以要加1
        # 然后除以7得到周数，由于是从周一开始算，所以需要加1
        week_number = delta.days // 7 + 1

        return week, section, week_number

    def get_week_section(used_str):
        # print(used_str)
        used_str = re.sub("\s", "/", used_str)
        if re.search("上午", used_str):
            # 4.1/周一上午
            used_str = re.sub("上午", "", used_str)
            used_str = re.sub("/", "/1-4/", used_str)
        elif re.search("下午", used_str):
            used_str = re.sub("下午", "", used_str)
            used_str = re.sub("/", "/5-8/", used_str)
        # print(used_str)
        return calculate_time(used_str)

    # 读取 data.xlsx 文件，加载其中的 “体检理论” 工作表，不含标题行
    df = pd.read_excel("data.xlsx", sheet_name=sheet, header=None)
    group = str(group)
    # 打印行数，列数
    # print(df.shape)

    # 逐行检查，如果一行内只有一个单元格横跨所有列，那么这一行是一个标题行，打印
    for i in range(df.shape[0]):
        if i == 0:
            subject = (
                df.iloc[i]
                .values[0]
                .replace("2021级北医预防医学专业", "")
                .replace("教学课程表", "")
            )
            print(subject)
        if df.iloc[i].count() == 1:
            # print(df.iloc[i].values[0])
            continue
        else:
            # 打印非标题行的内容
            print(df.iloc[i].values)
            # 获取其中的内容为 星期，节次，一线教师，授课地点，日期 所在的列索引
            # ['1、2组\n日期/节次' '3、4组\n日期/节次' '5、6组\n日期/节次' '7、8组\n日期/节次' '学时' '见习内容' '带教科室' '一线教师' '职称' '二线教师' '职称' '授课地点']
            # ['1组\n日期\n节次' '2组\n日期\n节次' '3组\n日期\n节次' '4组\n日期\n节次' '5组\n日期\n节次' '6组\n日期\n节次' '7组\n日期\n节次' '8组\n日期\n节次' '学时' '带教科室' '见习内容' '一线教师' '职称' '二线教师' '职称' '授课地点']
            used_index = 0
            for index, value in enumerate(df.iloc[i].values):
                if group in value:
                    used_index = index
                    break
            teacher_index = df.iloc[i].tolist().index("一线教师")
            content_index = df.iloc[i].tolist().index("见习内容")
            place_index = df.iloc[i].tolist().index("授课地点")
            department_index = df.iloc[i].tolist().index("带教科室")
            cur_index = i + 1
            break
    while True:
        is_exam = False
        # 检查是否又有标题行
        if df.iloc[cur_index].count() == 1:
            print(df.iloc[cur_index].values[0])
            break

        # 获取星期，节次，一线教师，授课地点，日期
        used_str = df.iloc[cur_index].tolist()[used_index]
        # 检查是否为 nan
        if used_str != used_str:
            used_str = df.iloc[cur_index].tolist()[0]
        if used_str != used_str:
            cur_index += 1
            continue
        week, section, week_number = get_week_section(used_str)
        content = df.iloc[cur_index].tolist()[content_index]

        # 检查是否为 nan
        if content != content:
            print("nan")
            is_exam = True
            content = df.iloc[cur_index].tolist()[department_index]
            print(content)
        elif "考试" in content:
            is_exam = True

        teacher = df.iloc[cur_index].tolist()[teacher_index]
        if teacher != teacher:
            teacher = "希波克拉底"
        place = df.iloc[cur_index].tolist()[place_index].replace("\n", "")

        if "-" in section:
            start_section, end_section = section.split("-")
        else:
            start_section = end_section = section

        # 写入到 csv 文件

        with open(output_file, "a") as f:
            name = subject
            if is_exam:
                name = content
            # 把 week 从 周一 周二 转换为 1 2
            week = week_to_number(week)
            teacher = teacher.replace("\n", "")
            f.write(
                f"{name},{week},{start_section},{end_section},{teacher},{place},{week_number}\n"
            )
        cur_index += 1


def main():
    # 读取 data.xlsx 文件，打印其中的所有工作表名
    wb = load_workbook("data.xlsx")
    sheets = wb.sheetnames
    # 移除 "课程总表"
    sheets.remove("课程总表")
    os.makedirs("./wakeup", exist_ok=True)
    print(sheets)
    for group in range(1, 9):
        # 先写入标题行
        output_file = f"./wakeup/第{group}组.csv"
        with open(output_file, "w") as f:
            f.write("课程名称,星期,开始节数,结束节数,老师,地点,周数\n")
        for sheet in sheets:
            if "见习" in sheet:
                convert_internship(sheet, group, output_file)
            else:
                convert_theroy(sheet, output_file)


if __name__ == "__main__":
    main()
